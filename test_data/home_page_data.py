import os

import openpyxl


class HomePageData:
    home_page_data = [
        {"name": "pepito", "mail": "pe@correo.com", "pass": "1234", "genre": "Male"},
        {"name": "pancha", "mail": "pa@correo.com", "pass": "BTC", "genre": "Female"}]

    @staticmethod
    def get_excel_data():
        package_dir = os.path.dirname(__file__)  # Obtiene el directorio del paquete actual
        excel_file_path = os.path.join(package_dir, "excel_data.xlsx") #Buscamos en el directorio actual el archivo excel

        data_excel = []  # creamos una lista para almacenar los diccionarios
        book = openpyxl.load_workbook(excel_file_path)  # busca el archivo excel
        sheet = book.active  # selecciona la hoja dentro del excel
        # obtener datos de un excel para guardarlos en un diccionario y posterior en una lista
        for i in range(1, sheet.max_row):
            dic = {}  # creamos un diccionario vacio el cual se limpia por cada fila recorrida
            for j in range(1, sheet.max_column - 1):
                if i != 1 and j != 1:
                    dic[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            if not dic: #si el diccionario esta vacio, continua para no guardar nada en el listado
                continue
            data_excel.append(dic)  # Almacenamos el diccionario obtenido dentro de la lista
        return data_excel
