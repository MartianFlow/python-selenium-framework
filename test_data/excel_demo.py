import openpyxl

book = openpyxl.load_workbook("excel_data.xlsx")  # busca el archivo excel
sheet = book.active  # selecciona la hoja dentro del excel
cell = sheet.cell(row=1, column=2)  # mapeo de la celda
print(cell.value)  # mapeo del valor obtenido de la celda

# Escribir en una celda en especifico
sheet.cell(row=2, column=2).value = "Jeffer"
print(sheet.cell(row=2, column=2).value)

print(f"Cantidad de filas maximas llenas: {sheet.max_row}")
print(f"Cantidad de columnas maximas llenas: {sheet.max_column}")
print(f"obtener un valor en especifico con referencia de celda: {sheet['A3'].value}")

# devolver todas las filas del excel
for i in range(1, sheet.max_row + 1):
    print(sheet.cell(row=i, column=1).value)

# devolver todas las columnas del excel a partir de una fila
for i in range(1, sheet.max_column + 1):
    print(sheet.cell(row=1, column=i).value)

print("#########################################")

#Mapeo de datos del excel a un diccionario en python
data_excel = [] #creamos una lista para almacenar los diccionarios

# matriz filas x columnas
for i in range(1, sheet.max_row):
    dict = {}  # creamos un diccionario vacio el cual se limpia por cada fila recorrida
    for j in range(1, sheet.max_column-1):
        if i != 1 and j != 1:
            dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

            #print(f"{sheet.cell(row=i, column=j).value}   ", end="")
        #print("")
    if not dict:
        continue
    data_excel.append(dict) #Almacenamos el diccionario obtenido dentro de la lista

print(data_excel)
